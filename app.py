import datetime
import json
import os
import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --- KONFIGURATION & DATEIPFEILE ---
EXCEL_FILE = "Gicht_Fitnees_APP.xlsx"
CACHE_FILE = "daily_cache.json"

st.set_page_config(
    page_title="Gicht & Fitness Tracker", page_icon="💪", layout="wide"
)

# --- AUTOMATISCHES ZWISCHENSPEICHERN & LADEN (JSON CACHE) ---
today_str = datetime.date.today().isoformat()


def load_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                # Prüfen, ob der Cache vom heutigen Tag ist
                if data.get("date") == today_str:
                    return data.get("state", {})
        except Exception:
            pass
    return None


def save_cache():
    # Alles aus st.session_state sichern, was für den Tag relevant ist
    state_to_save = {
        "date": today_str,
        "meals": st.session_state.get("meals", {}),
        "drinks": st.session_state.get("drinks", []),
        "workout": st.session_state.get("workout", {}),
        "daily_meta": st.session_state.get("daily_meta", {}),
    }
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(state_to_save, f, ensure_ascii=False, indent=4)
    except Exception:
        pass


# --- INITIALISIERUNG DES SESSION STATES ---
cached_state = load_cache()

if "meals" not in st.session_state:
    if cached_state and "meals" in cached_state:
        st.session_state["meals"] = cached_state["meals"]
    else:
        st.session_state["meals"] = {
            "Frühstück": [],
            "Mittagessen": [],
            "Abendessen": [],
            "Snacks": [],
        }

if "drinks" not in st.session_state:
    if cached_state and "drinks" in cached_state:
        st.session_state["drinks"] = cached_state["drinks"]
    else:
        st.session_state["drinks"] = []

if "workout" not in st.session_state:
    if cached_state and "workout" in cached_state:
        st.session_state["workout"] = cached_state["workout"]
    else:
        st.session_state["workout"] = {
            "done": False,
            "exercises": [],
            "duration": 0,
            "intensity": "Mittel",
        }

if "daily_meta" not in st.session_state:
    if cached_state and "daily_meta" in cached_state:
        st.session_state["daily_meta"] = cached_state["daily_meta"]
    else:
        st.session_state["daily_meta"] = {
            "weight": 75.0,
            "kfa": 15.0,
            "muscle": 60.0,
            "steps": 8000,
            "gicht_ampel": "Grün (Alles optimal)",
            "notes": "",
        }

# Jedes Mal bei Interaktion Cache aktualisieren
save_cache()

# --- SEITEN-LAYOUT & NAVIGATION ---
st.sidebar.title("Navigation")
menu = st.sidebar.selectbox(
    "Menü wählen",
    [
        "🏠 Dashboard & Übersicht",
        "🍳 Mahlzeiten & Ernährung",
        "💧 Wasser & Getränke",
        "🏋️‍♂️ Training & Aktivität",
        "⚖️ Körperwerte & Gicht-Ampel",
        "📊 Tagesabschluss & Excel-Export",
    ],
)

# --- 1. DASHBOARD ---
if menu == "🏠 Dashboard & Übersicht":
    st.title("💪 Gicht & Fitness Tages-Dashboard")
    st.write(
        f"Heutiges Datum: **{datetime.date.today().strftime('%d.%m.%Y')}** | Status: **Automatische Zwischenspeicherung aktiv 🟢**"
    )

    # Berechnungen
    total_kcal = 0
    total_protein = 0
    for category, items in st.session_state["meals"].items():
        for item in items:
            total_kcal += item.get("kcal", 0)
            total_protein += item.get("protein", 0)

    total_water = sum(
        d.get("amount", 0) for d in st.session_state["drinks"]
    )
    steps = st.session_state["daily_meta"].get("steps", 0)
    weight = st.session_state["daily_meta"].get("weight", 0)
    ampel = st.session_state["daily_meta"].get(
        "gicht_ampel", "Grün (Alles optimal)"
    )

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Kalorien gesamt", f"{total_kcal} kcal")
    col2.metric("Protein gesamt", f"{total_protein} g")
    col3.metric("Wasser / Trinken", f"{total_water} ml")
    col4.metric("Schritte", f"{steps}")

    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("Aktuelle Körperwerte")
        st.write(f"- **Gewicht:** {weight} kg")
        st.write(f"- **KFA:** {st.session_state['daily_meta'].get('kfa', 0)} %")
        st.write(
            f"- **Muskelmasse:** {st.session_state['daily_meta'].get('muscle', 0)} kg"
        )
    with c2:
        st.subheader("Gicht-Status & Wohlbefinden")
        st.info(f"**Aktuelle Ampel:** {ampel}")

    st.success(
        "💡 **Hinweis:** Du kannst die App jetzt jederzeit schließen. Deine Daten werden automatisch zwischengespeichert und beim nächsten Öffnen wiederhergestellt!"
    )

# --- 2. MAHLZEITEN & ERNÄHRUNG ---
elif menu == "🍳 Mahlzeiten & Ernährung":
    st.title("🍳 Mahlzeiten & Ernährung erfassen")

    meal_category = st.selectbox(
        "Mahlzeit", ["Frühstück", "Mittagessen", "Abendessen", "Snacks"]
    )

    with st.form("meal_form", clear_on_submit=True):
        food_name = st.text_input("Bezeichnung (z. B. Haferflocken mit Whey)")
        col_a, col_b = st.columns(2)
        with col_a:
            kcal_val = st.number_input(
                "Kalorien (kcal)", min_value=0, max_value=3000, value=300, step=10
            )
        with col_b:
            protein_val = st.number_input(
                "Protein (g)", min_value=0.0, max_value=300.0, value=25.0, step=1.0
            )
        submitted = st.form_submit_button("Mahlzeit hinzufügen")

        if submitted and food_name:
            st.session_state["meals"][meal_category].append(
                {"name": food_name, "kcal": kcal_val, "protein": protein_val}
            )
            save_cache()  # Sofort speichern
            st.success(f"'{food_name}' erfolgreich hinzugefügt!")
            st.rerun()

    st.subheader("Bisherige Einträge heute:")
    for cat, items in st.session_state["meals"].items():
        st.markdown(f"**{cat}**")
        if not items:
            st.write("Noch keine Einträge.")
        else:
            for idx, item in enumerate(items):
                col_i1, col_i2 = st.columns([4, 1])
                with col_i1:
                    st.write(
                        f"- {item['name']}: {item['kcal']} kcal, {item['protein']}g Protein"
                    )
                with col_i2:
                    if st.button("Löschen", key=f"del_meal_{cat}_{idx}"):
                        st.session_state["meals"][cat].pop(idx)
                        save_cache()
                        st.rerun()

# --- 3. WASSER & GETRÄNKE ---
elif menu == "💧 Wasser & Getränke":
    st.title("💧 Wasser- & Getränkemanagement (Gicht-Fokus)")

    with st.form("drink_form", clear_on_submit=True):
        drink_type = st.selectbox(
            "Art des Getränks",
            [
                "Wasser (still/mit Kohlensäure)",
                "Tee (ungesüßt)",
                "Kaffee",
                "Citruswasser / Zitronenwasser",
                "Sonstiges",
            ],
        )
        amount_ml = st.number_input(
            "Menge in ml", min_value=50, max_value=2000, value=500, step=50
        )
        drink_submitted = st.form_submit_button("Getränk eintragen")

        if drink_submitted:
            st.session_state["drinks"].append(
                {"type": drink_type, "amount": amount_ml}
            )
            save_cache()
            st.success(f"{amount_ml} ml {drink_type} eingetragen!")
            st.rerun()

    total_w = sum(d.get("amount", 0) for d in st.session_state["drinks"])
    st.metric("Gesamte Flüssigkeitsaufnahme heute", f"{total_w} ml")

    st.subheader("Getränke-Liste:")
    for idx, d in enumerate(st.session_state["drinks"]):
        col_d1, col_d2 = st.columns([4, 1])
        with col_d1:
            st.write(f"- {d['amount']} ml: {d['type']}")
        with col_d2:
            if st.button("Entfernen", key=f"del_drink_{idx}"):
                st.session_state["drinks"].pop(idx)
                save_cache()
                st.rerun()

# --- 4. TRAINING & AKTIVITÄT ---
elif menu == "🏋️‍♂️ Training & Aktivität":
    st.title("🏋️‍♂️ Training & Aktivität")

    with st.form("workout_form"):
        st.subheader("Workout-Daten")
        w_duration = st.number_input(
            "Dauer (Minuten)", min_value=0, max_value=300, value=45, step=5
        )
        w_intensity = st.selectbox("Intensität", ["Leicht", "Mittel", "Intensiv"])
        w_exercises = st.text_area(
            "Übungen / Notizen zum Training (z.B. Zirkeltraining, Hanteln)"
        )
        w_submit = st.form_submit_button("Training speichern")

        if w_submit:
            st.session_state["workout"] = {
                "done": True,
                "duration": w_duration,
                "intensity": w_intensity,
                "exercises": w_exercises,
            }
            save_cache()
            st.success("Training erfolgreich gespeichert!")

    st.subheader("Schritte erfassen")
    new_steps = st.number_input(
        "Tägliche Schritte",
        min_value=0,
        max_value=50000,
        value=int(st.session_state["daily_meta"].get("steps", 8000)),
        step=500,
    )
    if st.button("Schritte aktualisieren"):
        st.session_state["daily_meta"]["steps"] = new_steps
        save_cache()
        st.success("Schritte aktualisiert!")

# --- 5. KÖRPERWERTE & GICHT-AMPEL ---
elif menu == "⚖️ Körperwerte & Gicht-Ampel":
    st.title("⚖️ Körperwerte & Gicht-Symptome")

    w_val = st.number_input(
        "Körpergewicht (kg)",
        min_value=30.0,
        max_value=200.0,
        value=float(st.session_state["daily_meta"].get("weight", 75.0)),
        step=0.1,
    )
    kfa_val = st.number_input(
        "Körperfettanteil KFA (%)",
        min_value=3.0,
        max_value=50.0,
        value=float(st.session_state["daily_meta"].get("kfa", 15.0)),
        step=0.1,
    )
    mus_val = st.number_input(
        "Muskelmasse (kg)",
        min_value=10.0,
        max_value=150.0,
        value=float(st.session_state["daily_meta"].get("muscle", 60.0)),
        step=0.1,
    )

    ampel_val = st.selectbox(
        "Gicht-Ampel Status",
        [
            "Grün (Alles optimal)",
            "Gelb (Leichte Anzeichen / Vorsicht)",
            "Rot (Aktuelle Beschwerden / Gichtanfall)",
        ],
        index=0,
    )

    daily_notes = st.text_area(
        "Persönliche Notizen des Tages",
        value=st.session_state["daily_meta"].get("notes", ""),
    )

    if st.button("Werte speichern"):
        st.session_state["daily_meta"]["weight"] = w_val
        st.session_state["daily_meta"]["kfa"] = kfa_val
        st.session_state["daily_meta"]["muscle"] = mus_val
        st.session_state["daily_meta"]["gicht_ampel"] = ampel_val
        st.session_state["daily_meta"]["notes"] = daily_notes
        save_cache()
        st.success("Körperwerte und Gicht-Status erfolgreich gesichert!")

# --- 6. TAGESABSCHLUSS & EXCEL-EXPORT ---
elif menu == "📊 Tagesabschluss & Excel-Export":
    st.title("📊 Tagesabschluss & Excel-Export")
    st.write(
        "Hier siehst du die Zusammenfassung des Tages und kannst alle Daten final in die Excel-Tabelle schreiben."
    )

    # Berechnungen für Anzeige
    total_kcal = sum(
        i.get("kcal", 0)
        for cat in st.session_state["meals"].values()
        for i in cat
    )
    total_protein = sum(
        i.get("protein", 0)
        for cat in st.session_state["meals"].values()
        for i in cat
    )
    total_water = sum(
        d.get("amount", 0) for d in st.session_state["drinks"]
    )

    st.write(f"- **Kalorien:** {total_kcal} kcal")
    st.write(f"- **Protein:** {total_protein} g")
    st.write(f"- **Wasser:** {total_water} ml")
    st.write(
        f"- **Gewicht:** {st.session_state['daily_meta'].get('weight')} kg"
    )
    st.write(
        f"- **Gicht-Ampel:** {st.session_state['daily_meta'].get('gicht_ampel')}"
    )


    def save_current_day_to_excel():
        file_exists = os.path.exists(EXCEL_FILE)
        if file_exists:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            if "Tagebuch" in wb.sheetnames:
                ws = wb["Tagebuch"]
            else:
                ws = wb.active
                ws.title = "Tagebuch"
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tagebuch"
            # Header schreiben
            headers = [
                "Datum",
                "Gewicht (kg)",
                "KFA (%)",
                "Muskelmasse (kg)",
                "Schritte",
                "Kalorien (kcal)",
                "Protein (g)",
                "Wasser (ml)",
                "Gicht-Ampel",
                "Notizen",
            ]
            ws.append(headers)
            # Header styling
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="335577",
                    end_color="335577",
                    fill_type="solid",
                )
                cell.alignment = Alignment(
                    horizontal="center", vertical="center"
                )

        # Datensatz für heute anhängen
        row_data = [
            datetime.date.today().strftime("%Y-%m-%d"),
            st.session_state["daily_meta"].get("weight"),
            st.session_state["daily_meta"].get("kfa"),
            st.session_state["daily_meta"].get("muscle"),
            st.session_state["daily_meta"].get("steps"),
            total_kcal,
            total_protein,
            total_water,
            st.session_state["daily_meta"].get("gicht_ampel"),
            st.session_state["daily_meta"].get("notes"),
        ]
        ws.append(row_data)

        # Spaltenbreite anpassen
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(EXCEL_FILE)


    if st.button("In Excel speichern & Download vorbereiten"):
        save_current_day_to_excel()
        st.success("Tagesdaten erfolgreich in die Excel-Tabelle geschrieben!")

    if os.path.exists(EXCEL_FILE):
        with open(EXCEL_FILE, "rb") as f:
            st.download_button(
                label="📥 Excel-Tabelle herunterladen",
                data=f,
                file_name=EXCEL_FILE,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
